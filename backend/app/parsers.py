import csv
import io
from pathlib import Path
from typing import Callable, Dict

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


class UnsupportedFileTypeError(ValueError):
    pass


def _read_text_bytes(data: bytes) -> str:
    return data.decode('utf-8', errors='ignore')


def _parse_pdf(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages:
        text = page.extract_text() or ''
        if text:
            pages.append(text.strip())
    return '\n\n'.join(pages)


def _parse_docx(data: bytes) -> str:
    document = Document(io.BytesIO(data))
    paragraphs = [para.text.strip() for para in document.paragraphs if para.text.strip()]
    return '\n'.join(paragraphs)


def _parse_csv(data: bytes) -> str:
    text_stream = io.StringIO(_read_text_bytes(data))
    reader = csv.reader(text_stream)
    rows = ['\t'.join(cell.strip() for cell in row if cell.strip()) for row in reader]
    return '\n'.join(row for row in rows if row)


def _parse_xlsx(data: bytes) -> str:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        rows.append(f'# Sheet: {sheet.title}')
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                rows.append('\t'.join(cells))
    workbook.close()
    return '\n'.join(rows)


_PARSERS: Dict[str, Callable[[bytes], str]] = {
    '.txt': _read_text_bytes,
    '.md': _read_text_bytes,
    '.markdown': _read_text_bytes,
    '.pdf': _parse_pdf,
    '.docx': _parse_docx,
    '.csv': _parse_csv,
    '.xlsx': _parse_xlsx,
}


def extract_text(filename: str, data: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in _PARSERS:
        raise UnsupportedFileTypeError(f'Unsupported file type: {suffix}')
    return _PARSERS[suffix](data).strip()
